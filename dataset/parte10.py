import json

import pandas as pd
import re
from typing import List
from datetime import datetime


# Cargar los datos desde el archivo JSON
with open("DATOSJSON\datos_pacientes.json", "r", encoding="utf-8") as file:
    data = json.load(file)

# Lista para almacenar los datos de cada paciente
pacientes_data = []

# Diccionario de categorías de enfermedades
categorias = {
    "1. Enfermedad cardiovascular": [
        "cardiovascular", "infarto", "angina", "coronaria", "hta", "hipertensión arterial", "cardiopatía",
        "insuficiencia cardiaca", "arritmia", "taquicardia", "bradicardia", "miocardio", "ictus", "accidente cerebrovascular",
        "acv", "trombosis", "embolia", "hipertrofia ventricular"
    ],
    "2. Hipertensión arterial": [
        "hipertensión", "hipertension", "hta", "presión alta", "tensión alta", "crisis hipertensiva"
    ],
    "3. Diabetes mellitus": [
        "diabetes", "hiperglucemia", "glucosa alta", "glucemia elevada", "resistencia a la insulina"
    ],
    "4. Enfermedad renal": [
        "renal", "riñón", "insuficiencia renal", "nefropatía", "hemodiálisis", "diálisis", "litiasis renal",
        "cálculos renales", "proteinuria", "glomerulonefritis", "pielonefritis"
    ],
    "5. Enfermedad hepatica": [
        "hepática", "hígado", "cirrosis", "esteatosis", "hígado graso", "hepatitis", "transaminasas elevadas",
        "insuficiencia hepática", "colestasis", "fibrosis hepática"
    ],
    "6. Osteoporosis/ Artrosis/ Osteoartrosis": [
        "osteoporosis", "artrosis", "osteoartrosis", "coxartrosis", "gonartrosis", "desgaste articular",
        "dolor articular", "degeneración ósea", "fractura osteoporótica"
    ],
    "7. Enfermedad gastrointestinal": [
        "gastro", "colon", "gastritis", "úlceras", "gastrointestinal", "reflujo", "esofagitis", "colitis",
        "estreñimiento", "diarrea crónica", "síndrome de intestino irritable", "enfermedad de crohn",
        "rectocolitis", "pancreatitis", "helicobacter", "sucralfato"
    ],
    "8. Hipotiroidismo/Hipertiroidismo": [
        "hipotiroidismo", "hipertiroidismo", "tiroid", "tiroides", "bocio", "nódulo tiroideo", "hashimoto",
        "graves", "trastornos tiroideos"
    ],
    "9. Cancer": [
        "cáncer", "tumor", "neoplasia", "carcinoma", "adenocarcinoma", "linfoma", "melanoma", "leucemia",
        "sarcoma", "neoplasias", "cáncer de", "metástasis"
    ]
}

#clasificar por profesion----------------
def clasificar_escolaridad(escolaridad_raw):
    escolaridad_raw = str(escolaridad_raw).lower().strip()

    # Correcciones manuales comunes
    escolaridad_raw = escolaridad_raw.replace("ptofesional", "profesional")
    escolaridad_raw = escolaridad_raw.replace("bachillera", "bachillerato")

    # Eliminar entradas de ocupación
    if "ocupación" in escolaridad_raw:
        return 2  # Asumimos bachillerato

    if "analfabeta" in escolaridad_raw:
        return 4
    elif "primaria" in escolaridad_raw:
        return 3
    elif "bachiller" in escolaridad_raw:
        return 2
    elif "técnico" in escolaridad_raw or "tecnico" in escolaridad_raw:
        return 1
    elif "tecnólogo" in escolaridad_raw or "tecnologo" in escolaridad_raw:
        return 1
    elif "profesional" in escolaridad_raw:
        return 0
    elif "maestría" in escolaridad_raw or "maestria" in escolaridad_raw or "posgrado" in escolaridad_raw:
        return 0
    else:
        return 2  # Default conservador: bachillerato

#-----------------------------clasificar clinimetria-------------
def clasificar_clinimetria(tipo, valor):
    """
    Recibe el tipo de clinimetría ('das28', 'sledai', 'asdas') y su valor numérico,
    y devuelve una tupla con los valores correspondientes para cada columna.
    """
    das28, sledai, asdas = 0, 0, 0

    if tipo == "das28":
        if valor < 2.6:
            das28 = 1
        elif valor < 3.2:
            das28 = 2
        elif valor < 5.1:
            das28 = 3
        else:
            das28 = 4

    elif tipo == "sledai":
        if valor == 0:
            sledai = 0
        elif valor <= 5:
            sledai = 1
        elif valor <= 10:
            sledai = 2
        elif valor <= 19:
            sledai = 3
        else:
            sledai = 4

    elif tipo == "asdas":
        if valor < 1.3:
            asdas = 1
        elif valor < 2.1:
            asdas = 2
        elif valor < 3.5:
            asdas = 3
        else:
            asdas = 4
    else:
        sledai = 0
        asdas = 0
        das28 = 0

    return das28, sledai, asdas




# Tu función de extracción de fechas
import re
from typing import List

def extraer_fechas(texto: str, fecha_actual: str) -> List[str]:
    """
    Extrae fechas de un texto clínico, evitando confundirlas con dosis.
    Si el texto contiene la frase 'aún no iniciado' o 'aun no iniciado',
    añade la fecha_actual como indicativo de inicio del tratamiento.
    """

    # Patrones para detectar formatos válidos de fechas
    patrones = [
        r'\b(0[1-9]|[12][0-9]|3[01])[/-](0[1-9]|1[0-2])[/-](\d{2}|\d{4})\b',  # dd/mm/aa o dd-mm-aaaa
        r'\b(0[1-9]|1[0-2])[/-](\d{2}|\d{4})\b',                             # mm/aa o mm/aaaa
        r'\b(19|20)\d{2}\b',                                                # años: 2014, 2025
    ]

    # Patrones que deben excluirse (dosis como 1500/400 mg)
    patrones_exclusion = [
        r'\b\d{2,4}/\d{2,4}\s*(mg|ui|mg\/día|mg\/semana)?\b',
        r'\b\d{1,4}\s*(mg|ui|mg\/día|mg\/semana)\b',
    ]

    fechas_encontradas = []

    # Si el tratamiento no ha iniciado aún, se toma la fecha actual como fecha de inicio
    if re.search(r'\ba[uú]n\s+no(\s+ha)?\s+iniciado\b', texto, flags=re.IGNORECASE):
        fechas_encontradas.append(fecha_actual)

    # Buscar fechas reales y filtrar las que no son dosis
    for patron in patrones:
        for match in re.finditer(patron, texto, flags=re.IGNORECASE):
            posible_fecha = match.group()
            if not any(re.search(p_exc, posible_fecha, re.IGNORECASE) for p_exc in patrones_exclusion):
                fechas_encontradas.append(posible_fecha)

    return fechas_encontradas
def clasificar_fecha_tratamiento(fechas: List[str], fecha_actual_str: str) -> int:
    """
    Clasifica fechas en base a la más reciente:
    - 0 si no hay fechas válidas
    - 4 si la diferencia con la fecha actual es < 6 meses
    - 3 si la diferencia es entre 6 y < 12 meses
    - 1 si es 12 meses o más
    """

    if not fechas:
        return 0

    # Convertimos la fecha actual
    try:
        fecha_actual = datetime.strptime(fecha_actual_str, "%d-%m-%Y")
    except ValueError:
        print("⚠️ Fecha actual con formato incorrecto")
        return 0

    fechas_convertidas = []

    for fecha_str in fechas:
        fecha_str = fecha_str.strip()
        fecha = None

        # Intentamos distintos formatos
        formatos = [
            "%d/%m/%Y", "%d-%m-%Y",   # día completo
            "%d/%m/%y", "%d-%m-%y",   # día completo corto
            "%m/%Y", "%m-%Y",         # mes/año completo
            "%m/%y", "%m-%y",         # mes/año corto
            "%Y",                     # solo año completo
            "%y"                      # solo año corto
        ]

        for fmt in formatos:
            try:
                fecha = datetime.strptime(fecha_str, fmt)
                break
            except ValueError:
                continue

        # Si es solo año (ej: 25), lo interpretamos como 2025
        if fecha is None and re.match(r"^\d{2}$", fecha_str):
            year = int(fecha_str)
            year += 2000 if year < 50 else 1900
            fecha = datetime(year, 1, 1)

        if fecha:
            fechas_convertidas.append(fecha)

    if not fechas_convertidas:
        return 0

    # Usamos la fecha más reciente
    fecha_mas_reciente = max(fechas_convertidas)

    # Calcular diferencia en meses
    diferencia_meses = (fecha_actual.year - fecha_mas_reciente.year) * 12 + (fecha_actual.month - fecha_mas_reciente.month)

    # Clasificación según la diferencia
    if diferencia_meses < 6:
        return 4
    elif diferencia_meses < 12:
        return 3
    else:
        return 1

def evaluar_tratamiento_con_fecha_biologico_yak(texto: str, fecha_actual_str: str) -> int:
    medicamentos_interes = [
        "etanercept", "adalimumab", "rituximab", "tocilizumab", "abatacept",
        "infliximab", "golimumab", "certolizumab", "secukinumab", "belimumab",
        "tofacitinib", "baricitinib", "upadacitinib"
    ]

    # Unidades a ignorar si están presentes en el texto que parece fecha
    unidades_a_ignorar = ['mg', 'ui', 'vo', 'ml']

    if not isinstance(texto, str):
        return 0

    try:
        fecha_actual = datetime.strptime(fecha_actual_str, "%d-%m-%Y")
    except ValueError:
        print("⚠️ Fecha actual con formato incorrecto")
        return 0

    texto = texto.lower()
    tratamientos = re.split(r'[\n,*\-•]+', texto)
    tratamientos = [t.strip() for t in tratamientos if t.strip()]

    fechas_convertidas = []

    for item in tratamientos:
        for med in medicamentos_interes:
            if med in item:
                patrones = [
                    r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # dd/mm/yyyy o dd-mm-yyyy
                    r'\b\d{1,2}[/-]\d{2}',               # mm/yy
                    r'\b\d{4}\b',                        # Año completo
                    r'\b\d{2}\b'                         # Año corto
                ]
                fechas_encontradas = []
                for patron in patrones:
                    fechas_encontradas += re.findall(patron, item)

                for fecha_str in fechas_encontradas:
                    # Verificar si contiene unidades médicas
                    if any(unidad in fecha_str.lower() for unidad in unidades_a_ignorar):
                        continue

                    fecha_str = fecha_str.strip()
                    fecha = None
                    formatos = [
                        "%d/%m/%Y", "%d-%m-%Y",
                        "%d/%m/%y", "%d-%m-%y",
                        "%m/%Y", "%m-%Y",
                        "%m/%y", "%m-%y",
                        "%Y", "%y"
                    ]
                    for fmt in formatos:
                        try:
                            fecha = datetime.strptime(fecha_str, fmt)
                            break
                        except ValueError:
                            continue
                    if fecha is None and re.match(r"^\d{2}$", fecha_str):
                        year = int(fecha_str)
                        year += 2000 if year < 50 else 1900
                        fecha = datetime(year, 1, 1)
                    if fecha:
                        fechas_convertidas.append(fecha)

    if not fechas_convertidas:
        if any(med in item for item in tratamientos for med in medicamentos_interes):
            return 1
        else:
            return 0

    fecha_mas_reciente = max(fechas_convertidas)
    diferencia_meses = (fecha_actual.year - fecha_mas_reciente.year) * 12 + (fecha_actual.month - fecha_mas_reciente.month)

    if diferencia_meses < 3:
        return 4
    elif diferencia_meses < 6:
        return 3
    elif diferencia_meses < 12:
        return 1
    else:
        return 1



#evaluacio ntratamiendo dmards



def evaluar_tratamiento_con_fecha_dmards(texto: str, fecha_actual_str: str) -> int:
    medicamentos_interes = [
        "metotrexato", "leflunomida", "sulfasalazina", "hidroxicloroquina", "azatioprina",
        "ciclosporina", "ciclofosfamida", "micofenolato", "cloroquina"
    ]

    unidades_a_ignorar = ['mg', 'ui', 'vo', 'ml']

    if not isinstance(texto, str):
        return 0

    texto = texto.lower()

    try:
        fecha_actual = datetime.strptime(fecha_actual_str, "%d-%m-%Y")
    except ValueError:
        print("⚠️ Fecha actual con formato incorrecto")
        return 0

    tratamientos = re.split(r'[\n,*\-•]+', texto)
    tratamientos = [t.strip() for t in tratamientos if t.strip()]

    fechas_convertidas = []

    for item in tratamientos:
        for med in medicamentos_interes:
            if med in item:
                patrones = [
                    r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # 12/04/2023 o 12-04-2023
                    r'\b\d{1,2}[/-]\d{2}',               # 04/23
                    r'\b\d{4}\b',                        # Año completo
                    r'\b\d{2}\b'                         # Año corto (p. ej. 25)
                ]
                fechas_encontradas = []
                for patron in patrones:
                    fechas_encontradas += re.findall(patron, item)

                for fecha_str in fechas_encontradas:
                    if any(unidad in fecha_str.lower() for unidad in unidades_a_ignorar):
                        continue

                    fecha_str = fecha_str.strip()
                    fecha = None
                    formatos = [
                        "%d/%m/%Y", "%d-%m-%Y",
                        "%d/%m/%y", "%d-%m-%y",
                        "%m/%Y", "%m-%Y",
                        "%m/%y", "%m-%y",
                        "%Y", "%y"
                    ]
                    for fmt in formatos:
                        try:
                            fecha = datetime.strptime(fecha_str, fmt)
                            break
                        except ValueError:
                            continue
                    if fecha is None and re.match(r"^\d{2}$", fecha_str):
                        year = int(fecha_str)
                        year += 2000 if year < 50 else 1900
                        fecha = datetime(year, 1, 1)
                    if fecha:
                        fechas_convertidas.append(fecha)

    if not fechas_convertidas:
        if any(med in item for item in tratamientos for med in medicamentos_interes):
            return 1
        else:
            return 0

    fecha_mas_reciente = max(fechas_convertidas)
    diferencia_meses = (fecha_actual.year - fecha_mas_reciente.year) * 12 + (fecha_actual.month - fecha_mas_reciente.month)

    if diferencia_meses < 6:
        return 4
    elif diferencia_meses <12 :
        return 3
    elif (diferencia_meses > 12):
        return 1
    else:
        return 1

#-------------------------------------------------------------------
def evaluar_adherencia_biologicos(tratamiento_principal: str, adherencia_morisky: str) -> int:
    """
    Evalúa la adherencia a medicamentos biológicos según el tratamiento principal
    y el resultado del test de adherencia Morisky.

    Retorna:
        1 -> Adherente
        3 -> Parcialmente adherente
        4 -> No adherente
        0 -> No evaluable
    """

    # Lista de medicamentos biológicos en minúsculas
    biologicos = [
        "etanercept", "adalimumab", "rituximab", "tocilizumab", "abatacept",
        "infliximab", "golimumab", "certolizumab", "secukinumab", "belimumab"
    ]

    # Normalizar entradas
    tratamiento = str(tratamiento_principal).lower().replace("vita d", "vitamina d")
    adherencia = str(adherencia_morisky).strip().lower()

    # Verificar si hay al menos un biológico en el tratamiento
    usa_biologico = any(bio in tratamiento for bio in biologicos)

    # Evaluar adherencia
    if usa_biologico:
        if adherencia == "adherente":
            return 1
        elif adherencia == "parcialmente adherente":
            return 3
        elif adherencia == "no adherente":
            return 4
        else:
            return 0  # No evaluable
    else:
        return 0  # No usa biológicos, no se evalúa



#----------------------------------------------------------------------------
#adherencia antiyak
def evaluar_adherencia_inhibidoresjack(tratamiento_principal: str, adherencia_morisky: str) -> int:
    """
    Evalúa la adherencia a medicamentos biológicos según el tratamiento principal
    y el resultado del test de adherencia Morisky.

    Retorna:
        1 -> Adherente
        3 -> Parcialmente adherente
        4 -> No adherente
        0 -> No evaluable
    """

    # Lista de medicamentos jack en minúsculas
    inhibidores_jak = [
        "tofacitinib", "baricitinib", "upadacitinib"
    ]

    # Normalizar entradas
    tratamiento = str(tratamiento_principal).lower().replace("vita d", "vitamina d")
    adherencia = str(adherencia_morisky).strip().lower()

    # Verificar si hay al menos un jack en el tratamiento
    usa_jack = any(bio in tratamiento for bio in inhibidores_jak)

    # Evaluar adherencia
    if usa_jack:
        if adherencia == "adherente":
            return 1
        elif adherencia == "parcialmente adherente":
            return 3
        elif adherencia == "no adherente":
            return 4
        else:
            return 0  # No evaluable
    else:
        return 0  # No usa antijack, no se evalúa

#----------------------------------------------------
def evaluar_adherencia_demards(tratamiento_principal: str, adherencia_morisky: str) -> int:
    """
    Evalúa la adherencia a medicamentos biológicos según el tratamiento principal
    y el resultado del test de adherencia Morisky.

    Retorna:
        1 -> Adherente
        3 -> Parcialmente adherente
        4 -> No adherente
        0 -> No evaluable
    """

    # Lista de medicamentos jack en minúsculas
    dmards = [
        "leflunomida",
        "metotrexato",
        "cloroquina",
        "sulfasalazina",
        "prednisolona",
        "azatioprina",
        "hidroxicloroquina",
        "daflazacort",
        "etoricoxib"
    ]

    # Normalizar entradas
    tratamiento = str(tratamiento_principal).lower().replace("vita d", "vitamina d")
    adherencia = str(adherencia_morisky).strip().lower()

    # Verificar si hay al menos un jack en el tratamiento
    usa_demards = any(bio in tratamiento for bio in dmards)

    # Evaluar adherencia
    if usa_demards:
        if adherencia == "adherente":
            return 1
        elif adherencia == "parcialmente adherente":
            return 3
        elif adherencia == "no adherente":
            return 4
        else:
            return 0  # No evaluable
    else:
        return 0  # No usa demards, no se evalúa
#-------------------------- dispeesacion parenteral
def evaluar_parenteral():
    # Normalizamos el tratamiento principal
    tratamiento_principal = str(paciente_info.get("tratamiento_principal", "")).lower().replace("vita d", "vitamina d")

    # Lista de medicamentos de administración parenteral
    medicamentos_parenterales = [
        "metotrexato sc",
        "etanercept",
        "adalimumab",
        "rituximab",
        "tocilizumab",
        "abatacept",
        "infliximab",
        "golimumab",
        "certolizumab",
        "secukinumab",
        "belimumab"
    ]

    # Verificar si alguno de los medicamentos está en el tratamiento principal
    contiene_parenteral = any(med in tratamiento_principal for med in medicamentos_parenterales)

    # Si tiene medicamento parenteral, parenteral = 1, si no = 0
    return 1 if contiene_parenteral else 0


#----------------------------------------------------------------------
for paciente_id, paciente_info in data.items():
    if not isinstance(paciente_info, dict):
        continue

    print(paciente_info)

    nombre = str(paciente_info.get("nombre", "")).strip()

    # Edad y tipo identificación
    edad = paciente_info.get("edad")
    if isinstance(edad, int):
        tipo_identificacion = 1 if edad >= 18 else 2
    else:
        edad = None
        tipo_identificacion = None

    toma_escolaridad = str(paciente_info.get("nivel_escolaridad", "")).strip()
    escolaridad=clasificar_escolaridad(toma_escolaridad)
    # Género: 1 = femenino, 2 = masculino, 0 = desconocido
    observaciones = str(paciente_info.get("observaciones", "")).strip().upper()
    if "FEMENINO" in observaciones:
        genero = 1
    elif "MASCULINO" in observaciones:
        genero = 2
    else:
        genero = 1

    # Gestación: 0 para todos
    gestacion = 0

    # Consumo de SPA: verificamos alcohol, tabaco, sustancias
    alcohol = str(paciente_info.get("consumo_alcohol", "")).strip().strip('"').upper()
    tabaco = str(paciente_info.get("consumo_tabaco", "")).strip().strip('"').upper()
    sustancias = str(paciente_info.get("consumo_sustancias", "")).strip().strip('"').upper()

    if alcohol == "NIEGA" and tabaco == "NIEGA" and sustancias == "NIEGA":
        consumo_spa = 0
    elif tabaco != "NIEGA":
        consumo_spa = 1
    elif alcohol != "NIEGA":
        consumo_spa = 2
    elif sustancias != "NIEGA":
        consumo_spa = 3
    else:
        consumo_spa = 0

    # Columna vacía de interacción alcohol/drogas
    interaccion_medicamento =0

    # Trastornos mentales: 0 para todos
    trastornos_mentales = 0

    # Factores trato paciente: 0 para todos
    trato_paciente = 0

    # Hospitalización últimos 6 meses
    hospitalizacion_6_meses = str(paciente_info.get("hospitalizacion_ultimos_6_meses", "")).strip().lower()
    if hospitalizacion_6_meses in ["no", "no especificado"]:
        ultimos_6_meses = 0
    else:
        ultimos_6_meses = 4

    # Clasificación de enfermedades
    otro_diag = str(paciente_info.get("otro_diagnostico", "")).lower()
    clasificacion = {key: 0 for key in categorias}
    otros_enfermedades = []

    for enfermedad in re.split(r'[;,.\n]+', otro_diag):
        enfermedad = enfermedad.strip()
        clasificada = False
        for categoria, palabras in categorias.items():
            if any(p in enfermedad for p in palabras):
                clasificacion[categoria] = 1
                clasificada = True
                break
        if enfermedad and not clasificada and enfermedad != "niega":
            otros_enfermedades.append(enfermedad)

    # -------------------------
    # COMORBILIDADES (categorías del 1 al 9)
    comorbilidades = sum(clasificacion[c] for c in list(categorias.keys()))
    if comorbilidades == 1:
        presenta_comorbilidades = 2
    elif comorbilidades >= 2:
        presenta_comorbilidades = 4
    else:
        presenta_comorbilidades = 0

    # -------------------------
    # APLICA CLINIMETRÍA
    clinimetria_tipo = str(paciente_info.get("clinimetria_tipo", "")).strip().lower()
    if clinimetria_tipo and clinimetria_tipo != "no aplica":
        aplica_clinimetria = 4
    else:
        aplica_clinimetria = 0
    # -------------------------
    # COMORBILIDADES (categorías del 1 al 9)
    comorbilidades = sum(clasificacion[c] for c in list(categorias.keys()))
    if comorbilidades == 1:
        presenta_comorbilidades = 2
    elif comorbilidades >= 2:
        presenta_comorbilidades = 4
    else:
        presenta_comorbilidades = 0

    # -------------------------
    # APLICA CLINIMETRÍA
    clinimetria_tipo = str(paciente_info.get("clinimetria_tipo", "")).strip().lower()
    if clinimetria_tipo and clinimetria_tipo != "no aplica":
        aplica_clinimetria = 4
    else:
        aplica_clinimetria = 0

    # === CLASIFICACIÓN DE CLINIMETRÍA ===
    clinimetria_tipo = str(paciente_info.get("clinimetria_tipo", "")).strip().lower()
    clinimetria_valor = paciente_info.get("clinimetria_valor", None)

    # Normalizar el tipo de clinimetría
    if "das28" in clinimetria_tipo:
        clinimetria_tipo = "das28"
    elif "asdas" in clinimetria_tipo:
        clinimetria_tipo = "asdas"
    elif "sledai" in clinimetria_tipo:
        clinimetria_tipo = "sledai"
    else:
        clinimetria_tipo = ""

    # Validar si el valor es un número
    try:
        clinimetria_valor = float(clinimetria_valor)
    except (ValueError, TypeError):
        clinimetria_valor = None

    # Clasificar sólo si tenemos tipo válido y valor numérico
    if clinimetria_tipo and clinimetria_valor is not None:
        das28_clasificacion, sledai_clasificacion, asdas_clasificacion = clasificar_clinimetria(clinimetria_tipo,
                                                                                                clinimetria_valor)
    else:
        das28_clasificacion = sledai_clasificacion = asdas_clasificacion = 0

    # ------------------------------------
    # CAMBIO EN LA MEDICACIÓN (columna nueva)

    # CAMBIO EN LA MEDICACIÓN (columna nueva)

    # Extraer campos de tratamiento
    tratamiento_principal = str(paciente_info.get("tratamiento_principal", "")).lower().replace("vita d", "vitamina d")
    conciliacion_meds = str(paciente_info.get("conciliacion_medicamentos", "")).lower().replace("vita d", "vitamina d")

    # Unir textos
    todos_meds = f"{tratamiento_principal}\n{conciliacion_meds}"

    # Separar solo por saltos de línea y comas (indicadores de medicamentos distintos)

    meds_separados = re.split(r'[\n,]', todos_meds)

    # Limpiar y mantener medicamentos que contienen '+', ya que son compuestos
    lista_meds = [med.strip() for med in meds_separados if med.strip()]
    print("MEDICAMENTOS:", lista_meds)

    # Extraer el nombre base del medicamento (primeras dos palabras por seguridad)
    nombres_base = set()
    for med in lista_meds:
        palabras = med.split()
        if len(palabras) >= 2:
            nombre_base = " ".join(palabras[:2])
        elif palabras:
            nombre_base = palabras[0]
        else:
            continue
        nombres_base.add(nombre_base)

    # Evaluar cuántos medicamentos únicos hay
    polimedicacion = 4 if len(nombres_base) >= 5 else 0

    #-------------------------- mirar la fecha de los medicamentos
    # Extraer campos de tratamiento
    tratamiento_principal = str(paciente_info.get("tratamiento_principal", "")).lower().replace("vita d", "vitamina d")
    fecha_actual=str(paciente_info.get("fecha_consulta", "")).lower()
    # Extraer fechas del texto de tratamiento
    fechas_tratamiento = extraer_fechas(tratamiento_principal,fecha_actual)
    print("Fechas encontradas:", fechas_tratamiento)

    cambio_en_medicacion=clasificar_fecha_tratamiento(fechas_tratamiento,fecha_actual)
#--------------------------------------------------------------------------

    #EVALUAR INICIO BIOLOGICO YACK DMARDS ----------------------------------------------

    tratamiento_principal = str(paciente_info.get("tratamiento_principal", "")).lower().replace("vita d", "vitamina d")
    inicio_tratamiendo_biologico_yack=evaluar_tratamiento_con_fecha_biologico_yak(tratamiento_principal,fecha_actual)
    #EVALUAR DEMARDS
    inicio_tratamiendo_demards=evaluar_tratamiento_con_fecha_dmards(tratamiento_principal,fecha_actual)
# adeherencia test miroski biologico ------------------------------------------------------------------
    tratamiento_principal =str(paciente_info.get("tratamiento_principal", "")).lower().replace("vita d", "vitamina d")
    adherencia_mirosky=str(paciente_info.get("adherencia_global","")).lower()
    BIOLOGICO_ADHERENCIA_MIROSKY = evaluar_adherencia_biologicos(tratamiento_principal, adherencia_mirosky)
#ADHERENCIA inibidores ajck
    JACK_ADHERENCIA_MIROSKY=evaluar_adherencia_inhibidoresjack(tratamiento_principal,adherencia_mirosky)
#ADHERENCIA DMARDS
    DMARDS_ADHERENCIA_MIROSKY=evaluar_adherencia_demards(tratamiento_principal,adherencia_mirosky)
#---------columan -adherencia a otros tratamiento -----------------------
    adherencia_mirosky = str(paciente_info.get("adherencia_global", "")).lower()

    # Evaluar adherencia para otros tratamientos
    if adherencia_mirosky == "no adherente":
        adherencia_otros_tratamientos = 4
    elif adherencia_mirosky == "parcialmente adherente":
        adherencia_otros_tratamientos = 3
    elif adherencia_mirosky == "adherente":
        adherencia_otros_tratamientos = 1
    else:
        adherencia_otros_tratamientos = 0  # No evaluable

#---------------------------------DISPENSACION tengo 2 columasn aca la columan  ORAL y la columan  parenteral---------------------------------
    adherencia_mirosky = str(paciente_info.get("adherencia_global", "")).lower()
    dispensacion = str(paciente_info.get("dispensacion", "")).lower()

    # Inicializamos el valor
    valor_dispensacion_oral = 0
    parenteral = 0
    # Solo entrar si la adherencia es "adherente" o "parcialmente adherente"
    if adherencia_mirosky in ["adherente", "parcialmente adherente"]:
        if "dispensacion parcial" in dispensacion:
            valor_dispensacion_oral = 2
            parenteral=evaluar_parenteral()
        elif "dispensacion completa" in dispensacion:
            valor_dispensacion_oral = 1
            parenteral=0
        elif "no identificado" in dispensacion:
            valor_dispensacion_oral = 0
            parenteral = 0
        elif "no dispensacion" in dispensacion:
            valor_dispensacion_oral = 3
            parenteral = 0
        else:
            valor_dispensacion_oral = 0  # Puedes poner -1 para casos no contemplados
            parenteral = 0

#--------------------- INTERACCIOENS SI NO ---------------------------
    texto = str(paciente_info.get("interacciones_farmacologicas", "")).lower()

    # Lista de medicamentos clave en minúsculas
    medicamentos_clave = [
        "leflunomida", "metotrexato", "cloroquina", "sulfasalazina", "prednisolona",
        "azatioprina", "hidroxicloroquina", "daflazacort", "etoricoxib", "etanercept",
        "adalimumab", "rituximab", "tocilizumab", "abatacept", "infliximab",
        "golimumab", "certolizumab", "secukinumab", "belimumab"
    ]

    # Buscar texto después de "significativas:" o "significativa:"
    match = re.search(r'significativas?:\s*(.*)', texto)
    contenido = match.group(1).strip() if match else ""

    if "ninguna" in contenido or contenido == "":
        interacciones = 0
        interacciones_mayores = 0
        clasificacion_relevancia = 0
        mecanismo = 0
        descripcion_molecula = 0
    else:
        interacciones = 1
        # Filtrar medicamentos presentes en el contenido
        descripcion_molecula = "; ".join([
    med.capitalize() for med in medicamentos_clave if med in contenido
    ]) or 0

        interacciones_mayores = 2
        clasificacion_relevancia = 3
        mecanismo = 1 if descripcion_molecula else 2
    #--------------------colimna de tipo farmaco siemroe es 1
    tipo_farmaco=1
    #-.--------------------------- RAM
    texto=str(paciente_info.get("ram")).lower()
    ram = 0 if "niega" in texto else 1
    ram_0no_1si=0 if "niega" in texto else 1
    # Nuevas columnas RAM adicionales
    columnas_ram_adicionales = {
        "MEDICAMENTO_SOSPECHOSO": 0 if ram == 0 else "",
        "ESTADO ACUTAL RAM": 0 if ram == 0 else "",
        "0NO 2 SI": 0 if ram == 0 else "",
        "TIPO": 0 if ram == 0 else "",
        "DEFINIDO POR COMITE FV": 0 if ram == 0 else "",
        "0 NO 1 SI": 0 if ram == 0 else "",
        "ADMINISTRACION ERRONEA DEL MEDICAMENTO": 0 if ram == 0 else "",
        "CARACTERISTICAS PERSONALES": 0 if ram == 0 else "",
        "CONVERSACION INADECUADA": 0 if ram == 0 else "",
        "CONTRAINDICACION": 0 if ram == 0 else "",
        "DOSIS PAUTAS": 0 if ram == 0 else "",
        "DUPLICIDAD": 0 if ram == 0 else "",
        "ERRROES EN LA DISPENSACION": 0 if ram == 0 else "",
        "ERRORES EN LA PRESCIRPCION": 0 if ram == 0 else "",
        "IN CUMPLIMIENTO": 0 if ram == 0 else "",
        "INTERACCIONES": 0 if ram == 0 else "",
        "OTROS PROBLEMAS": 0 if ram == 0 else "",
        "PROBAVBILIDAD DE EFECTOS ADVERSOS": 0 if ram == 0 else "",
        "PROBLEMA DE SALUD": 0 if ram == 0 else "",
        "OTROS": 0 if ram == 0 else ""
    }



#-------------------------------------------------------------------------------------

    # Guardar los datos del paciente
    fila_paciente = {
        "Nombre": nombre,
        "Tipo Identificación": tipo_identificacion,
        "Edad": edad,
        "Grado Escolaridad": escolaridad,
        "Género": genero,
        "Gestación": gestacion,
        "Consumo de SPA": consumo_spa,
        "4. Consumo de alcohol o drogas que interacciona con medicamento": interaccion_medicamento,
        "Trastornos mentales": trastornos_mentales,
        "Factores relacionados con el trato paciente": trato_paciente,
        "hospitalizacion ultimos 6 meses": ultimos_6_meses,
        # Enfermedades (se asignan desde el diccionario `clasificacion`)
        "1. Enfermedad cardiovascular": clasificacion["1. Enfermedad cardiovascular"],
        "2. Hipertensión arterial": clasificacion["2. Hipertensión arterial"],
        "3. Diabetes mellitus": clasificacion["3. Diabetes mellitus"],
        "4. Enfermedad renal": clasificacion["4. Enfermedad renal"],
        "5. Enfermedad hepatica": clasificacion["5. Enfermedad hepatica"],
        "6. Osteoporosis/ Artrosis/ Osteoartrosis": clasificacion["6. Osteoporosis/ Artrosis/ Osteoartrosis"],
        "7. Enfermedad gastrointestinal": clasificacion["7. Enfermedad gastrointestinal"],
        "8. Hipotiroidismo/Hipertiroidismo": clasificacion["8. Hipotiroidismo/Hipertiroidismo"],
        "9. Cancer": clasificacion["9. Cancer"],
        "10. Otros": 1 if otros_enfermedades else 0,
        "¿Cuáles otras?": ", ".join(otros_enfermedades)if otros_enfermedades else 0,
        "4. Presenta más de 2 comorbilidades de la lista \n2. Presenta 1 comorbilidad de la lista": presenta_comorbilidades,
        "APLICA CLINIMETRÍA\n0. No, requiere de otro parametro\n4. Si, pero es > a 2 meses": aplica_clinimetria,
        "das28_clasificacion": das28_clasificacion,
        "sledai_clasificacion": sledai_clasificacion,
        "asdas_clasificacion": asdas_clasificacion,
        "polimedicacion": polimedicacion,
        "Cambio en Medicacion":cambio_en_medicacion,
        "INICIO TRATAMIENTO  BIOLOGICO / ANTIYACK": inicio_tratamiendo_biologico_yack,
        "INICIO TRATAMIENTO DMARDS":inicio_tratamiendo_demards,
        "ADHERENCIA MIROSKY GREEN BIOLOGICO":BIOLOGICO_ADHERENCIA_MIROSKY,
        "ADHERENCIA MIROSKY GREEN JACK":JACK_ADHERENCIA_MIROSKY,
        "ADHERENCIA MIROSKY DMARDS":DMARDS_ADHERENCIA_MIROSKY,
        "ADHERENCIA A OTROS TRATAMIENTOS FARMACOLOGICOS":adherencia_otros_tratamientos,
        "Dispensacion parenteral":parenteral,
        "Dispesacion medicamentos oral": valor_dispensacion_oral,
        "Interacciones 1si 2no":interacciones,
        "interacciones mayores que requieran":interacciones_mayores,
        "clasificacion relevancia":clasificacion_relevancia,
        "mecanismo farmadinamicas":mecanismo,
        "farmaco":tipo_farmaco,
        "descripcion de las molecula":descripcion_molecula,
        "columna 0no 1 si de ram":ram_0no_1si,
        "RAM": ram,
        **columnas_ram_adicionales

    }

    # Agregar columnas de clasificación
    fila_paciente.update(clasificacion)

    # Añadir al conjunto de datos
    pacientes_data.append(fila_paciente)


# Crear DataFrame
df = pd.DataFrame(pacientes_data)

# Ordenar las columnas

# Concatenamos en el orden deseado: otras columnas, luego las de clasificación por enfermedades, y al final las de clinimetría
columnas_finales = ["Nombre"]+["Tipo Identificación"]+["Edad"]+["Grado Escolaridad"]+["Género"]+["Gestación"]+["Consumo de SPA"]+["4. Consumo de alcohol o drogas que interacciona con medicamento"]+["Trastornos mentales"]+["Factores relacionados con el trato paciente"]+["hospitalizacion ultimos 6 meses"]+["1. Enfermedad cardiovascular"]+["2. Hipertensión arterial"]+["3. Diabetes mellitus"]+["4. Enfermedad renal"]+["5. Enfermedad hepatica"]+["6. Osteoporosis/ Artrosis/ Osteoartrosis"]+["7. Enfermedad gastrointestinal"]+["8. Hipotiroidismo/Hipertiroidismo"]+["9. Cancer"]+["10. Otros"]+["¿Cuáles otras?"]+["4. Presenta más de 2 comorbilidades de la lista \n2. Presenta 1 comorbilidad de la lista"]+["APLICA CLINIMETRÍA\n0. No, requiere de otro parametro\n4. Si, pero es > a 2 meses"]+["das28_clasificacion"]+["sledai_clasificacion"]+["asdas_clasificacion"] +["polimedicacion"]+["Cambio en Medicacion"]+["INICIO TRATAMIENTO  BIOLOGICO / ANTIYACK"] + ["INICIO TRATAMIENTO DMARDS"]+["ADHERENCIA MIROSKY GREEN BIOLOGICO"]+["ADHERENCIA MIROSKY GREEN JACK"]+["ADHERENCIA MIROSKY DMARDS"]+["ADHERENCIA A OTROS TRATAMIENTOS FARMACOLOGICOS"]+["Dispensacion parenteral"]+["Dispesacion medicamentos oral"]+["Interacciones 1si 2no"]+["interacciones mayores que requieran"]+["clasificacion relevancia"]+["mecanismo farmadinamicas"]+["farmaco"]+["descripcion de las molecula"]+["columna 0no 1 si de ram"]+["RAM"]+[  # Aquí agregas todas las columnas nuevas
        "MEDICAMENTO_SOSPECHOSO",
        "ESTADO ACUTAL RAM",
        "0NO 2 SI",
        "TIPO",
        "DEFINIDO POR COMITE FV",
        "0 NO 1 SI",
        "ADMINISTRACION ERRONEA DEL MEDICAMENTO",
        "CARACTERISTICAS PERSONALES",
        "CONVERSACION INADECUADA",
        "CONTRAINDICACION",
        "DOSIS PAUTAS",
        "DUPLICIDAD",
        "ERRROES EN LA DISPENSACION",
        "ERRORES EN LA PRESCIRPCION",
        "IN CUMPLIMIENTO",
        "INTERACCIONES",
        "OTROS PROBLEMAS",
        "PROBAVBILIDAD DE EFECTOS ADVERSOS",
        "PROBLEMA DE SALUD",
        "OTROS"
    ]

# Reordenar el DataFrame según el orden definido
df = df[columnas_finales]

# Guardar a Excel
df.to_excel("HOJAexcelSUBIR\pacientes_detallado.xlsx", index=False)
# ------------------------------------------
# Aplicar formato rojo si RAM == 1
# ------------------------------------------
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Abrir el archivo recién guardado
archivo = "HOJAexcelSUBIR\pacientes_detallado.xlsx"
libro = load_workbook(archivo)
hoja = libro.active

# Buscar la columna "RAM"
columna_ram = None
for col in range(1, hoja.max_column + 1):
    if hoja.cell(row=1, column=col).value == "RAM":
        columna_ram = col
        break

# Aplicar color rojo si RAM == 1
if columna_ram:
    for fila in range(2, hoja.max_row + 1):
        celda = hoja.cell(row=fila, column=columna_ram)
        if celda.value == 1:
            celda.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Guardar cambios
libro.save(archivo)

